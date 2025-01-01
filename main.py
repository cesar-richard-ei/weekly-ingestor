from datetime import datetime, timedelta
import openpyxl

input_file_path = "./export2.xlsx"

workbook = openpyxl.load_workbook(input_file_path)
sheet = workbook.active

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

print(f"Number of rows: {sheet.max_row}")
data_by_date = {}

for row in sheet.iter_rows(min_row=2, values_only=True):
    client, project, hour_date, hour_tags, hour_note = row
    if hour_date and hour_note:
        date_obj = datetime.strptime(hour_date, "%d/%m/%Y")
        if date_obj not in data_by_date:
            data_by_date[date_obj] = []
        prefix = f"[{project}]" if project in ["CI", "DevOps"] else ""
        data_by_date[date_obj].append((prefix, hour_note))

sorted_dates = sorted(data_by_date.keys())

# Ajouter des dates manquantes
if sorted_dates:
    start_date = sorted_dates[0]
    end_date = sorted_dates[-1]
    current_date = start_date
    while current_date <= end_date:
        if current_date not in data_by_date:
            data_by_date[current_date] = []
        current_date += timedelta(days=1)

sorted_dates = sorted(data_by_date.keys())

for index, date in enumerate(sorted_dates, start=1):
    notes = []
    for prefix, note in data_by_date[date]:
        note_lines = note.split("\n")
        note_lines[0] = f"{prefix} {note_lines[0]}" if prefix else f"{note_lines[0]}"
        note = "\n".join(note_lines)
        notes.append(note)
    cell_value = "\n\n".join(notes)
    cell_date = new_sheet.cell(row=index, column=1)
    cell_time = new_sheet.cell(row=index, column=2)
    cell_client = new_sheet.cell(row=index, column=3)
    cell_location = new_sheet.cell(row=index, column=4)
    cell_tasks = new_sheet.cell(row=index, column=5)

    cell_date.value = date.strftime("%d/%m/%Y")
    if notes:
        if len(notes) == 1 and notes[0].strip() == "OFF":
            cell_time.value = "0"
            cell_client.value = ""
            cell_location.value = ""
            cell_tasks.value = cell_value
        elif "OFF" in notes:
            cell_time.value = "0.5"
            cell_client.value = "Pasqal"
            cell_location.value = "Remote"
            cell_tasks.value = cell_value
        else:
            cell_time.value = "1"
            cell_client.value = "Pasqal"
            cell_location.value = "Remote"
            cell_tasks.value = cell_value
    else:
        cell_time.value = "0"
        cell_client.value = ""
        cell_location.value = ""
        cell_tasks.value = ""

print(f"Number of rows: {len(sorted_dates)}")

# Sauvegarder le nouveau classeur
output_file_path = "./imputations.xlsx"
new_workbook.save(output_file_path)
