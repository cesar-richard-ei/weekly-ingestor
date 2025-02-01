"""
Script de génération de rapports d'activité à partir des événements Timely.
Génère un fichier Excel formaté selon les besoins spécifiques.
"""

from datetime import datetime, timedelta
import httpx
import asyncio
import os
from dotenv import load_dotenv
import openpyxl
from typing import Dict, List, Tuple
import holidays

# Configuration
load_dotenv()
TIMELY_ACCOUNT_ID = os.getenv("TIMELY_ACCOUNT_ID")
API_BASE_URL = "http://localhost:8000"

# Initialisation des jours fériés français
fr_holidays = holidays.FR()

class TimelyReport:
    """Gestionnaire de rapports Timely"""

    def __init__(self, account_id: str, api_url: str):
        self.account_id = account_id
        self.api_url = api_url

    async def get_events(self, from_date: str, to_date: str) -> List[Dict]:
        """Récupère tous les événements pour une période donnée"""
        async with httpx.AsyncClient() as client:
            all_events = []
            page = 1

            while True:
                response = await client.get(
                    f"{self.api_url}/{self.account_id}/events",
                    params={
                        "since": from_date,
                        "upto": to_date,
                        "page": page,
                        "per_page": 250,
                    },
                )
                events = response.json()
                if not events:
                    break

                all_events.extend(events)
                page += 1

            return all_events

    def filter_events_by_client(self, events: List[Dict], client_name: str) -> List[Dict]:
        """Filtre les événements pour ne garder que ceux du client spécifié"""
        return [
            event for event in events
            if event.get("project", {}).get("client", {}).get("name", "") == client_name
        ]

    @staticmethod
    def _is_weekend(date: datetime) -> bool:
        """Vérifie si une date est un weekend"""
        return date.weekday() >= 5

    @staticmethod
    def _is_holiday(date: datetime) -> bool:
        """Vérifie si une date est un jour férié"""
        return date in fr_holidays

    @staticmethod
    def _get_dates_range(from_date: str, to_date: str) -> List[datetime]:
        """Génère la liste des dates dans l'intervalle"""
        start = datetime.strptime(from_date, "%Y-%m-%d")
        end = datetime.strptime(to_date, "%Y-%m-%d")
        delta = (end - start).days + 1
        return [start + timedelta(days=x) for x in range(delta)]

    def process_events(
        self, events: List[Dict], from_date: str, to_date: str
    ) -> Dict[datetime, List[Tuple[str, str]]]:
        """Traite les événements et les organise par date"""
        data_by_date = {}

        # Initialiser toutes les dates
        for date in self._get_dates_range(from_date, to_date):
            if self._is_weekend(date):
                data_by_date[date] = [("", "WEEKEND")]
            elif self._is_holiday(date):
                data_by_date[date] = [("", "HOLIDAY")]
            else:
                data_by_date[date] = []

        # Ajouter les événements
        for event in events:
            date = datetime.strptime(event["day"], "%Y-%m-%d")
            if not (self._is_weekend(date) or self._is_holiday(date)):
                project = event.get("project", {}).get("name", "")
                is_special = project in ["CI", "DevOps"]
                prefix = f"[{project}]" if is_special else ""
                data_by_date[date].append((prefix, event.get("note", "")))

        return data_by_date

    def generate_excel(
        self, data_by_date: Dict[datetime, List[Tuple[str, str]]], output_path: str
    ):
        """Génère le fichier Excel à partir des données traitées"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for index, (date, entries) in enumerate(sorted(data_by_date.items()), 1):
            self._write_row(sheet, index, date, entries)

        workbook.save(output_path)

    def _write_row(
        self, sheet, row: int, date: datetime, entries: List[Tuple[str, str]]
    ):
        """Écrit une ligne dans le fichier Excel"""
        notes = []
        for prefix, note in entries:
            if note in ["WEEKEND", "HOLIDAY"]:
                notes = [note]
                break
            note_lines = note.split("\n")
            if prefix:
                note_lines[0] = f"{prefix} {note_lines[0]}"
            notes.append("\n".join(note_lines))

        cell_value = "\n\n".join(notes)

        # Remplir les cellules
        sheet.cell(row=row, column=1, value=date.strftime("%d/%m/%Y"))

        if notes:
            if notes[0] in ["WEEKEND", "HOLIDAY"]:
                time, client, location = "0", "", ""
            elif len(notes) == 1 and notes[0].strip() == "OFF":
                time, client, location = "0", "", ""
            elif any(note.strip() == "OFF" for note in notes):
                time, client, location = "0.5", "Pasqal", "Remote"
            else:
                time, client, location = "1", "Pasqal", "Remote"
        else:
            time, client, location = "0", "", ""

        sheet.cell(row=row, column=2, value=time)
        sheet.cell(row=row, column=3, value=client)
        sheet.cell(row=row, column=4, value=location)
        sheet.cell(row=row, column=5, value=cell_value)


async def main():
    """Point d'entrée principal"""
    # Calculer la période (mois courant)
    today = datetime.now()
    from_date = today.replace(day=1).strftime("%Y-%m-%d")
    to_date = (
        today.replace(
            day=1,
            month=today.month + 1 if today.month < 12 else 1,
            year=today.year if today.month < 12 else today.year + 1,
        )
        - timedelta(days=1)
    ).strftime("%Y-%m-%d")

    # Générer le rapport
    report = TimelyReport(TIMELY_ACCOUNT_ID, API_BASE_URL)
    events = await report.get_events(from_date, to_date)
    data = report.process_events(events, from_date, to_date)
    report.generate_excel(data, "./imputations.xlsx")
    print(f"Rapport généré pour la période du {from_date} au {to_date}")


if __name__ == "__main__":
    asyncio.run(main())
